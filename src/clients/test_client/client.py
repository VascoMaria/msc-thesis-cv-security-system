import cv2
import asyncio
import aiohttp
import argparse
import threading
import time
from datetime import datetime
import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook, load_workbook
import os
from asyncio import TimeoutError as AsyncTimeout

# --- Excel logging setup ------------------------------------------
ENABLE_EXCEL_LOG = False   # muda para False se não quiseres guardar
EXCEL_PATH = "tempos_processamentoSystem.xlsx"
sheet_name = datetime.now().strftime("Exec_%Y-%m-%d_%H-%M-%S")

if not os.path.exists(EXCEL_PATH):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(["Processing Time (s)"])
    wb.save(EXCEL_PATH)
else:
    wb = load_workbook(EXCEL_PATH)
    ws = wb.create_sheet(title=sheet_name)
    ws.append(["Processing Time (s)"])
    wb.save(EXCEL_PATH)


class App(tk.Tk):
    def __init__(self, api_port):
        super().__init__()
        self.title("Security System Client")
        self.geometry("900x500")
        self.protocol("WM_DELETE_WINDOW", self.on_close)

        self.API_URL    = f"http://localhost:{api_port}/process_frame"
        self.STATUS_URL = f"http://localhost:{api_port}/status"

        self.streaming = False
        self.fps_var   = tk.StringVar(value="10")

        # remember last logged batch key
        self._last_batch_key = None

        self.build_ui()

        # background asyncio loop
        self.loop = asyncio.new_event_loop()
        threading.Thread(target=self.loop.run_forever, daemon=True).start()

    def build_ui(self):
        left = ttk.Frame(self, width=240, padding=10)
        left.pack(side="left", fill="y")

        ttk.Button(left, text="Start Streaming", command=self.start).pack(fill="x")
        ttk.Button(left, text="Stop Streaming",  command=self.stop).pack(fill="x", pady=(5,20))

        ttk.Label(left, text="FPS:").pack(anchor="w")
        ttk.Entry(left, textvariable=self.fps_var).pack(fill="x", pady=(0,10))

        ttk.Button(left, text="Get Status", command=self.get_status).pack(fill="x")
        self.status_output = ttk.Label(
            left,
            text="No status yet.",
            wraplength=220,
            justify="left",
            background="#f0f0f0",
            relief="sunken",
            padding=5
        )
        self.status_output.pack(fill="x", pady=(5,0))

        right = ttk.Frame(self)
        right.pack(side="right", fill="both", expand=True)

        notebook = ttk.Notebook(right)
        notebook.pack(fill="both", expand=True)

        # All Logs tab
        frame_all = ttk.Frame(notebook)
        self.log_all = tk.Text(
            frame_all, bg="black", fg="white", state="disabled", font=("Consolas", 10)
        )
        self.log_all.pack(fill="both", expand=True)
        notebook.add(frame_all, text="All Logs")

        # POSTs tab
        frame_posts = ttk.Frame(notebook)
        self.log_posts = tk.Text(
            frame_posts, bg="black", fg="lightgreen", state="disabled", font=("Consolas", 10)
        )
        self.log_posts.pack(fill="both", expand=True)
        notebook.add(frame_posts, text="POSTs")

    def log_message(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_all.configure(state="normal")
        self.log_all.insert("end", f"{ts}  {msg}\n")
        self.log_all.see("end")
        self.log_all.configure(state="disabled")

    def log_post(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_posts.configure(state="normal")
        self.log_posts.insert("end", f"{ts}  {msg}\n")
        self.log_posts.see("end")
        self.log_posts.configure(state="disabled")

    def start(self):
        if self.streaming:
            return
        try:
            fps = float(self.fps_var.get())
            assert fps > 0
        except:
            self.log_message("❌ Invalid FPS — enter a positive number.")
            return

        self.streaming = True
        self.log_message(f"▶️  Starting stream at {fps:.1f} FPS")
        threading.Thread(target=self.capture_loop, daemon=True).start()

    def stop(self):
        if not self.streaming:
            return
        self.streaming = False
        self.log_message("⏹ Stopped stream")

    def get_status(self):
        async def _():
            async with aiohttp.ClientSession() as sess:
                try:
                    r = await sess.get(self.STATUS_URL, timeout=5)
                    js = await r.json()
                except Exception as e:
                    self.log_message(f"Status error: {e}")
                    return

            # format timestamp
            ts = js.get("timestamp", "")
            if ts:
                ts = ts.split(".")[0].replace("T", " ")

            if js.get("message") == "No alarms have occurred yet.":
                text = "✅ No alarms have occurred yet."
            elif js.get("message") == "Alarm detected":
                dets = js.get("detections", [])
                dets_str = ", ".join(dets) if dets else "none"
                text = f"🚨 Last alarm at:\n{ts}\nDetections: {dets_str}"
            else:
                text = js.get("message", str(js))

            self.status_output.config(text=text)
            self.log_message("👉 Status updated")

        asyncio.run_coroutine_threadsafe(_(), self.loop)

    def capture_loop(self):
        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            self.log_message("❌ Cannot open camera")
            return

        try:
            fps = float(self.fps_var.get())
        except:
            fps = 10.0
        interval = 1.0 / fps

        session = aiohttp.ClientSession(loop=self.loop)

        async def send_frame(frame):
            _, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 80])
            form = aiohttp.FormData()
            form.add_field("frame", buf.tobytes(),
                           filename="frame.jpg", content_type="image/jpeg")

            # log the POST being sent
            self.log_post(f"POST /process_frame  size={len(buf)} bytes")

            try:
                resp = await session.post(self.API_URL, data=form, timeout=30)
                rj   = await resp.json()

                # build a simple de-dup key: (rounded time, tuple of detections)
                t    = rj.get("processing_time", 0.0)
                dets = tuple(rj.get("detections", []))
                extra_info = rj.get("extra_info", [])
                key  = (round(t, 3), dets)

                # only log when this batch key changes
                if key != self._last_batch_key:
                    self._last_batch_key = key
                    alarm = rj.get("alarm", False)
                    self.log_message(f"→ alarm={alarm} time={t:.2f}s dets={list(dets)} extra_info={extra_info}")
                    # Excel logging
                    wb = load_workbook(EXCEL_PATH)
                    ws = wb[sheet_name]
                    ws.append([t])
                    wb.save(EXCEL_PATH)

            except AsyncTimeout:
                return
            except OSError:
                return
            except Exception as e:
                self.log_message(f"⚠️ Unexpected send error: {e}")

        while self.streaming and cap.isOpened():
            start = time.time()
            ret, frame = cap.read()
            if not ret:
                self.log_message("❌ Capture error")
                break

            cv2.imshow("Streaming to API", frame)
            # fire-and-forget the POST
            asyncio.run_coroutine_threadsafe(send_frame(frame), self.loop)

            elapsed = time.time() - start
            time.sleep(max(0, interval - elapsed))

            if cv2.waitKey(1) & 0xFF == ord("q"):
                self.streaming = False
                break

        cap.release()
        cv2.destroyAllWindows()
        asyncio.run_coroutine_threadsafe(session.close(), self.loop)
        self.log_message("🔌 Stream loop ended")

    def on_close(self):
        self.stop()
        self.destroy()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="GUI Frame Sender Client")
    parser.add_argument("--port", "-p", type=int, default=8050,
                        help="FastAPI port (e.g. 8050)")
    args = parser.parse_args()

    App(api_port=args.port).mainloop()
