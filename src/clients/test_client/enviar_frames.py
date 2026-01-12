import cv2
import asyncio
import aiohttp
import argparse


import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

EXCEL_PATH = "tempos_processamentoSystem.xlsx"
sheet_name = datetime.now().strftime("Exec_%Y-%m-%d_%H-%M-%S")

# Cria o Excel e a nova aba se necessário
if not os.path.exists(EXCEL_PATH):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(["Tempo de Processamento (s)"])
    wb.save(EXCEL_PATH)
else:
    wb = load_workbook(EXCEL_PATH)
    ws = wb.create_sheet(title=sheet_name)
    ws.append(["Tempo de Processamento (s)"])
    wb.save(EXCEL_PATH)



# URL da API onde os frames serão enviados
API_URL: str

# FPS máximo permitido (pode ser ajustado)
MAX_FPS: float = 10.0  # Define um limite de 10 frames por segundo

async def send_frame(session, frame):
    """ Converte e envia um frame comprimido para a API de forma assíncrona """
    _, encoded_image = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 80])  # Compressão para reduzir tamanho
    frame_bytes = encoded_image.tobytes()

    form_data = aiohttp.FormData()
    form_data.add_field("frame", frame_bytes, filename="frame.jpg", content_type="image/jpeg")
    #print(f"Tamanho do frame (bytes): {len(frame_bytes)}")

    try:
        async with session.post(API_URL, data=form_data, timeout=30) as response:
            response_data = await response.json()
            print(f"Resposta da API: {response_data}")  # Apenas para debug
            # Se o campo de tempo existir, salvar no Excel
            if response_data.get("status") == "discarded":
                print(response_data.get("message", "Frame descartado"))
                return  # Não tenta usar processing_time

            if "processing_time" in response_data:
                try:
                    processing_time = float(response_data["processing_time"])
                    wb = load_workbook(EXCEL_PATH)
                    ws = wb[sheet_name]
                    ws.append([processing_time])
                    wb.save(EXCEL_PATH)
                except Exception as e:
                    print(f"Erro ao salvar tempo no Excel: {e}")
    except Exception as e:
        return

async def capture_and_send():
    """ Captura frames da câmera e os envia continuamente para a API """
    cap = cv2.VideoCapture(0)  # Captura da câmera (0 = webcam padrão)
    width = cap.get(cv2.CAP_PROP_FRAME_WIDTH)
    height = cap.get(cv2.CAP_PROP_FRAME_HEIGHT)
    print(f"Resolução da câmera: {int(width)}x{int(height)}")
    if not cap.isOpened():
        print("Erro ao abrir a câmera!")
        return

    async with aiohttp.ClientSession() as session:
        while True:
            start_time = asyncio.get_event_loop().time()  # Marca o tempo inicial

            ret, frame = cap.read()
            if not ret:
                print("Erro ao capturar frame!")
                continue

            # Exibir frame na tela sem bloquear o envio
            cv2.imshow("Streaming para API", frame)

            # Envia frame para a API sem bloquear a captura da câmera
            asyncio.create_task(send_frame(session, frame))

            # Ajusta taxa de envio para respeitar o FPS máximo
            elapsed_time = asyncio.get_event_loop().time() - start_time
            
            sleep_time = max(0, (1 / MAX_FPS) - elapsed_time)
            await asyncio.sleep(sleep_time)

            
            if cv2.waitKey(1) & 0xFF == ord("q"):
                break

    cap.release()  # Libera a câmera quando o loop termina
    cv2.destroyAllWindows()  # Fecha a janela de vídeo

# Inicia a captura e envio de frames de forma assíncrona
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Captura e envia frames ao endpoint /process_frame"
    )
    parser.add_argument(
        "--port", "-p",
        type=int,
        default=8050,
        help="Porta onde a API FastAPI está escutando (ex: 8050)"
    )
    parser.add_argument(
        "--max-fps", "-f",
        type=float,
        default=10.0,
        help="Máximo de frames por segundo"
    )
    args = parser.parse_args()

    # Ajusta a URL da API e o MAX_FPS de acordo com argumentos
    API_URL = f"http://localhost:{args.port}/process_frame"
    MAX_FPS = args.max_fps if args.max_fps > 0 else float('inf')

    print(f"Client connected to {API_URL} — limite de FPS: {args.max_fps}")
    asyncio.run(capture_and_send())

