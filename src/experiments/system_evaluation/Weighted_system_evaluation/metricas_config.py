import pandas as pd
import os
from openpyxl import load_workbook
from sklearn.metrics import confusion_matrix, accuracy_score, recall_score, precision_score, f1_score

# Caminho do Excel de input e output
input_path = "resultado_avaliacao_datasetValidacao.xlsx"
output_path = "metricas_avaliacao_datasetValidacao.xlsx"

# Carrega os dados
df = pd.read_excel(input_path)
df["true"] = df["Label do dataset"].map({"Positive": 1, "Negative": 0})
df["pred"] = df["Resultado do sistema"].map({True: 1, False: 0})

# Verifica se há erros nos dados
if df["true"].isnull().any() or df["pred"].isnull().any():
    raise ValueError("❌ Há valores inválidos no Excel.")

# Confusion matrix
tn, fp, fn, tp = confusion_matrix(df["true"], df["pred"]).ravel()

# Métricas
metrics = {
    "True Positives (TP)": tp,
    "True Negatives (TN)": tn,
    "False Positives (FP)": fp,
    "False Negatives (FN)": fn,
    "Accuracy": round(accuracy_score(df["true"], df["pred"]), 4),
    "Precision": round(precision_score(df["true"], df["pred"]), 4),
    "Recall (Sensibilidade)": round(recall_score(df["true"], df["pred"]), 4),
    "F1-score": round(f1_score(df["true"], df["pred"]), 4),
    "False Negative Rate (FNR)": round(fn / (fn + tp), 4) if (fn + tp) > 0 else 0
}
metrics_df = pd.DataFrame(list(metrics.items()), columns=["Métrica", "Valor"])

# Escrever no Excel
if not os.path.exists(output_path):
    # Se não existir, cria novo
    metrics_df.to_excel(output_path, index=False, sheet_name="Sheet1")
else:
    # Se já existir, abre e cria nova sheet
    book = load_workbook(output_path)
    sheet_count = len(book.sheetnames)
    new_sheet_name = f"Sheet{sheet_count + 1}"

    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
        metrics_df.to_excel(writer, index=False, sheet_name=new_sheet_name)


print(f"✅ Métricas gravadas em '{output_path}' na sheet '{new_sheet_name if os.path.exists(output_path) else 'Sheet1'}'.")
